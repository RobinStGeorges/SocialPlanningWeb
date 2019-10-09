from flask import Flask, send_file, Response, send_from_directory
from flask import redirect, request, url_for, render_template
from flask_wtf import Form
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from flask import Flask, session
from flask_session.__init__ import Session
from oauthlib.oauth2 import WebApplicationClient
from wtforms import StringField, TextField, HiddenField
from wtforms.validators import DataRequired
from openpyxl import Workbook
import notification
from user import User
from flask_mysqldb import MySQL
from oauth2client import client
from oauth2client import tools
from httplib2 import Http
from oauth2client import file
from oauth2client.client import flow_from_clientsecrets
from urllib.parse import urlparse
from googleEvent import GoogleEvent
from flask import Flask, render_template
from flask_wtf import Form
from flask import Flask, render_template
from flask_wtf import Form
from wtforms import Form
from dateutil.parser import parse as dtparse
from datetime import datetime as dt
from wtforms import StringField, DateTimeField,  BooleanField

import os
import os.path
import requests
import random
import sqlite3
import string
import pickle
import calendar
import json
import pickle
import google.oauth2.credentials
import google.oauth2.credentials
import google_auth_oauthlib.flow
import oauth2client
import sys
import datetime
import user
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)

# --------CONFIGURATION--------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY") or os.urandom(24)

SCOPES = 'https://www.googleapis.com/auth/calendar'

app.config['MYSQL_HOST'] = 'Kireta.mysql.pythonanywhere-services.com'
app.config['MYSQL_USER'] = 'Kireta'
app.config['MYSQL_PASSWORD'] = 'fulljuju'
app.config['MYSQL_DB'] = 'Kireta$socialPlanning'
mysql = MySQL(app)

GOOGLE_CLIENT_ID = '359467559574-l7s2bhh1kh6a2qagck59nvieco7868jo.apps.googleusercontent.com'
GOOGLE_CLIENT_SECRET = 'vnZgm3owuNgjsHWa-MtqpN_E'
GOOGLE_DISCOVERY_URL = (
    "https://accounts.google.com/.well-known/openid-configuration"
)

client = WebApplicationClient(GOOGLE_CLIENT_ID)

THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
FILE_URL =  os.path.join(THIS_FOLDER, 'socialPlanning.apk')


# --------controles--------#
login_manager = LoginManager()
login_manager.init_app(app)


@login_manager.user_loader
def load_user(user_id):
    return User.get(mysql, user_id)


@login_manager.unauthorized_handler
def unauthorized():
    return "You must be logged in to access this content.", 403


# --------fonctions--------#

def getCurrentUserName():
    return current_user.name + ' ' + current_user.familly_name


def get_google_provider_cfg():
    return requests.get(GOOGLE_DISCOVERY_URL).json()


def getAllEventsObjects():
    service = getGoogleCalendarService()
    now = datetime.datetime.utcnow().isoformat() + 'Z'  # 'Z' indicates UTC time
    events_result = service.events().list(calendarId='primary', timeMin=now,
                                          maxResults=10, singleEvents=True,
                                          orderBy='startTime').execute()
    events = events_result.get('items', [])
    if not events:
        print('No upcoming events found.')
    else:
        return events


def getGoogleEvents(maxEventsNumber):
    service = getGoogleCalendarService()
    now = datetime.datetime.utcnow().isoformat() + 'Z'  # 'Z' indicates UTC time
    events_result = service.events().list(calendarId='primary', timeMin=now,
                                          maxResults=maxEventsNumber, singleEvents=True,
                                          orderBy='startTime').execute()
    events = events_result.get('items', [])
    return events


def getGoogleEventsFilterDate(maxEventsNumber, date):
    service = getGoogleCalendarService()
    now = datetime.datetime.utcnow().isoformat() + 'Z'  # 'Z' indicates UTC time
    if date == '':
        dateFIso = now
    else:
        dateF = (datetime.datetime.strptime(date, '%d-%m-%Y')) - datetime.timedelta(days=1)
        dateFIso = dateF.isoformat() + ".649730Z"
    events_result = service.events().list(calendarId='primary', timeMin=dateFIso,
                                          maxResults=maxEventsNumber, singleEvents=True,
                                          orderBy='startTime').execute()
    events = events_result.get('items', [])
    return events


def saveGoogleEventsToDatabase(events):
    for event in events:
        try:
            event['summary']
        except NameError:
            print("No title")
        else:
            mail = event['organizer']['email']
            # start = event['start']['dateTime']
            start = event['start'].get('dateTime', event['start'].get('date'))
            dateStartParsed = start[0:9]
            # end =event['end']['dateTime']
            end = event['end'].get('dateTime', event['end'].get('date'))
            dateEndParsed = end[0:9]
            titre = event['summary']
            idGoogle = event["id"]
            GoogleEvent.create(mysql, idGoogle, mail, start, end, titre)


def getGoogleCalendarService():
    creds = None
    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    clientSecret = os.path.join(THIS_FOLDER, 'client_secrets.json')
    tokenFile = os.path.join(THIS_FOLDER, current_user.tokenPath)


    if os.path.exists(tokenFile):
        with open(tokenFile, 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                clientSecret, SCOPES)
            creds = flow.run_local_server()
        with open(tokenFile, 'wb') as token:
            pickle.dump(creds, token)
    service = build('calendar', 'v3', credentials=creds)
    return service


def getAllEvents():
    service = getGoogleCalendarService()

    now = datetime.datetime.utcnow().isoformat() + 'Z'  # 'Z' indicates UTC time
    events_result = service.events().list(calendarId='primary', timeMin=now,
                                          maxResults=100, singleEvents=True,
                                          orderBy='startTime').execute()
    events = events_result.get('items', [])
    json = "["
    if not events:
        print('No upcoming events found.')
    i = 0
    for event in events:
        if i != 0:
            json += ","
        start = event['start'].get('dateTime', event['start'].get('date'))
        dateParsed = formatDate(start)
        json += "{\"title\": \"" + event['summary'] + "\" , \"start\": \"" + dateParsed + "\", \"id\": \"" + event['id'] + "\"}"
        i += 1
    json += "]"
    return json


def formatDate(date):
    return date[0:10] + ' ' + date[11:13] + ':' + date[14:16] + ':00'


def isSameDay(dateSource, dateCible):
    yearDate = dateSource[-4:]
    monthDate = dateSource[3:5]
    dayDate = dateSource[0:2]

    yearCible = dateCible[0:4]
    monthCible = dateCible[5:7]
    dayCible = dateCible[8:10]

    if (yearDate == yearCible and monthDate == monthCible and dayDate == dayCible):
        return True
    return False


def filterEventsByDate(date, events):
    if not events:
        return []
    listFilteredEvents = []
    for event in events:
        start = event[1].strftime("%Y-%m-%d %H:%M:%S")
        end = event[2].strftime("%Y-%m-%d %H:%M:%S")
        isSameDayStart = isSameDay(date, start)
        isSameDayEnd = isSameDay(date, end)
        if (isSameDayStart or isSameDayEnd):
            listFilteredEvents.append(event)
    return listFilteredEvents


def getListStartEnd(event):
    if isinstance(event, list):
        event = event[0]
    listStartEnd = []
    mail = event[0]
    start = event[1].strftime("%Y-%m-%d %H:%M:%S")
    end = event[2].strftime("%Y-%m-%d %H:%M:%S")
    listStartEnd.append({'mail': mail, 'start': start, 'end': end})
    print("getListStartEnd liststartend")
    print(listStartEnd)
    return listStartEnd

def createNotification(email, content, idEvent, accepted):
    notification.create(mysql, email, content, idEvent, accepted)

# --------form--------
class eventForm(Form):
    title = StringField('titre', validators=[DataRequired()])
    description = StringField('description', validators=[DataRequired()])
    start = DateTimeField('start', format='%Y-%m-%d %H:%M:%S')
    end = DateTimeField('end', format='%Y-%m-%d %H:%M:%S')
    typeEvent = HiddenField("typeEvent")


class dateFilter(Form):
    dateFilterEvents = StringField('dateFilterEvents')


# --------ROUTES--------

# route vérification google ( à refaire, changement d'hebergeur)
@app.route("/google5e89b3351265473e.html")
def verifgoogle():
    return render_template('google5e89b3351265473e.html')


@app.route("/")
def index():
    if current_user.is_authenticated:
        if 'oldName' in session:
            if (session['oldName'] != getCurrentUserName()):
                session['oldName'] = getCurrentUserName()
                if (current_user.tokenPath != None):
                    if os.path.exists(current_user.tokenPath):
                        os.remove(current_user.tokenPath)
        else:
            session['oldName'] = getCurrentUserName()
        #return render_template('index.html', test='', name=current_user.name + ' ' + current_user.familly_name.upper())
        return redirect(url_for("calendrier"))
    else:
        return render_template('login.html')


@app.route("/login")
def login():
    google_provider_cfg = get_google_provider_cfg()
    authorization_endpoint = google_provider_cfg["authorization_endpoint"]
    request_uri = client.prepare_request_uri(
        authorization_endpoint,
        redirect_uri=request.base_url + "/callback",
        scope=["openid", "email", "profile"],
    )
    return redirect(request_uri + '&prompt=consent')


@app.route("/login/callback")
def callback():
    code = request.args.get("code")
    google_provider_cfg = get_google_provider_cfg()
    token_endpoint = google_provider_cfg["token_endpoint"]
    token_url, headers, body = client.prepare_token_request(
        token_endpoint,
        authorization_response=request.url,
        redirect_url=request.base_url,
        code=code,
    )
    token_response = requests.post(
        token_url,
        headers=headers,
        data=body,
        auth=(GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET),
    )
    client.parse_request_body_response(json.dumps(token_response.json()))
    userinfo_endpoint = google_provider_cfg["userinfo_endpoint"]
    uri, headers, body = client.add_token(userinfo_endpoint)
    userinfo_response = requests.get(uri, headers=headers, data=body)
    if userinfo_response.json().get("email_verified"):
        unique_id = userinfo_response.json()["sub"]
        users_email = userinfo_response.json()["email"]
        picture = userinfo_response.json()["picture"]
        users_name = userinfo_response.json()["given_name"]
        users_familly_name = userinfo_response.json()["family_name"]
        users_locale = userinfo_response.json()["locale"]
    else:
        return "User email not available or not verified by Google.", 400
    user = User(
        id_=unique_id, name=users_name, email=users_email, profile_pic=picture, familly_name=users_familly_name,
        locale=users_locale
    )
    if not User.get(mysql, unique_id):
        User.create(mysql, unique_id, users_name, users_email, picture, users_familly_name, users_locale)
    login_user(user)
    return redirect(url_for("index"))


@app.route("/calendrier")
def calendrier():
    # delete all events that were before in the database
    GoogleEvent.clean(mysql, current_user.email)

    # get x events from user's google calendar account
    events = getGoogleEvents(100)

    # put updated events into the database
    if events == "Bad Request":
        print('No upcoming events found.')
    else:
        saveGoogleEventsToDatabase(events)

    # Create javascript calender
    tc = calendar.HTMLCalendar(firstweekday=0)
    year = datetime.datetime.now().year

    return render_template('calendrier.html', cal=tc.formatyear(year, 3), username=session['oldName'],
                           name=current_user.name + ' ' + current_user.familly_name.upper())


@app.route('/events', methods=["GET", "POST"])
def showEvents():
    form = dateFilter()
    if request.method == "POST":
        date = request.form.get('dateFilterEvents')
        # get x events from user's google calendar account
        events = getGoogleEventsFilterDate(100, date)
        return render_template('events.html', events=events,
                               name=current_user.name + ' ' + current_user.familly_name.upper(), form=form)
    else:
        # delete all events that were before in the database
        GoogleEvent.clean(mysql, current_user.email)

        # get x events from user's google calendar account
        events = getGoogleEvents(100)

        # put updated events into the database
        if events == "Bad Request":
            print('No upcoming events found.')
        else:
            saveGoogleEventsToDatabase(events)
        events = getAllEventsObjects()
        return render_template('events.html', events=events,
                               name=current_user.name + ' ' + current_user.familly_name.upper(), form=form)


@app.route("/userInfo")
def userInfo():
    print(FILE_URL)
    return render_template('userInfo.html', userName=current_user.name, email=current_user.email,
                           profilPic=current_user.profile_pic, fName=current_user.familly_name,
                           locale=current_user.locale, name=current_user.name + ' ' + current_user.familly_name.upper(), fileUrl = FILE_URL)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))


@app.route('/data')
def return_data():
    return getAllEvents()


@app.route('/getUsersOccupiedHours', methods=["GET", "POST"])
def getUsersOccupiedHours():
    data = request.get_json()
    mails = data['mails']
    date = data['date']
    print("getUsersOccupiedHours, mails, date")
    print(mails)
    print(date)
    dictStartEndByEmail = []
    eventFilteredByDate = []
    for email in mails:
        print("email, eventsFiltered")
        events = GoogleEvent.get(mysql, email)
        if not eventFilteredByDate:
            print("if not")
            eventFilteredByDate = filterEventsByDate(date, events)
        else:
            print("else")
            eventFilteredByDate.append(filterEventsByDate(date, events))
    print(eventFilteredByDate)
    for event in eventFilteredByDate:
        dictStartEndByEmail.append(getListStartEnd(event))
    print("dictStartEndByEmail")
    print(dictStartEndByEmail)
    return json.dumps(dictStartEndByEmail)


@app.route('/addEvent', methods=["GET", "POST"])
def addEvent():
    form = eventForm()
    if request.method == "POST":
        title = request.form.get('title')
        description = request.form.get('description')
        start = request.form.get('start')
        end = request.form.get('end')
        datetimeStart = datetime.datetime.strptime(start, '%d-%m-%Y %H:%M').isoformat()
        datetimeEnd = datetime.datetime.strptime(end, '%d-%m-%Y %H:%M').isoformat()
        sizeForm = len(request.form)
        attendees = []
        if (sizeForm >= 6):
            for i in range(5, sizeForm):
                dictTemp = {}
                index = i - 4
                dictTemp['email'] = request.form['mail' + str(index)]
                attendees.append(dictTemp)
        service = getGoogleCalendarService()
        event = {
            'summary': title,
            'description': description,
            'start': {
                'dateTime': datetimeStart,
                'timeZone': 'Europe/Paris',
            },
            'end': {
                'dateTime': datetimeEnd,
                'timeZone': 'Europe/Paris',
            },
            'attendees': attendees,
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email', 'minutes': 24 * 60},
                    {'method': 'popup', 'minutes': 10},
                ],
            },
        }
        event = service.events().insert(calendarId='primary', body=event).execute()
        for aMail in attendees:
            notification.create(mysql, aMail['email'], 'invitation a l\'evenement '+event["summary"]+" de "+current_user.email, event["id"], "en attente de réponse")
        return redirect(url_for("showEvents"))
    return render_template('add.html', form=form, name=current_user.name + ' ' + current_user.familly_name.upper())


@app.route('/delete/<idGoogle>')
def delete(idGoogle):
    service = getGoogleCalendarService()
    service.events().delete(calendarId='primary', eventId=idGoogle).execute()
    return redirect(url_for("showEvents"))


@app.route('/download')
def download():
    email = current_user.email
    datas = GoogleEvent.get(mysql, email)

    workbook = Workbook()
    sheet = workbook.active

    # header
    sheet.cell(row=1, column=1).value = "Mail"
    sheet.cell(row=1, column=2).value = "start"
    sheet.cell(row=1, column=3).value = "end"
    sheet.cell(row=1, column=4).value = "nom evenement"
    sheet.cell(row=1, column=5).value = "id evenement"

    # content
    rowCurrent = 2
    for data in datas:
        sheet.cell(row=rowCurrent, column=1).value = data[0]
        sheet.cell(row=rowCurrent, column=2).value = data[1].strftime("%m/%d/%Y %H:%M:%S")
        sheet.cell(row=rowCurrent, column=3).value = data[2].strftime("%m/%d/%Y %H:%M:%S")
        sheet.cell(row=rowCurrent, column=4).value = data[3]
        sheet.cell(row=rowCurrent, column=5).value = data[5]
        rowCurrent = rowCurrent + 1

    THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
    filenameXlsx = os.path.join(THIS_FOLDER, email + 'downloadedDatas.xlsx')
    workbook.save(filename=filenameXlsx)

    return send_from_directory(directory=THIS_FOLDER, filename=email + 'downloadedDatas.xlsx', as_attachment=True)


@app.route('/deleteDatas')
def deleteDatas():
    GoogleEvent.clean(mysql, email=current_user.email)
    user.delete(mysql, email=current_user.email)
    logout_user()
    return redirect(url_for("index"))


@app.route('/showEvent/<id>/')
def showEvent(id):
    service = getGoogleCalendarService()
    event = service.events().get(calendarId='primary', eventId=id).execute()
    return render_template('event.html', event=event, name=current_user.name + ' ' + current_user.familly_name.upper())

@app.route('/notifs')
def showNotifs():
    notifs = notification.get(mysql, current_user.email)
    return render_template('notifications.html', notifications = notifs, name=current_user.name + ' ' + current_user.familly_name.upper())

@app.route('/notif/<id>')
def showNotif(id):
    notif = notification.getById(id)
    return render_template('notification.html', notification = notif, name=current_user.name + ' ' + current_user.familly_name.upper(), email = current_user.email)

@app.route('/manageNotificationResponse/<id>/<reponse>/<idNotif>')
def manageNotificationResponse(id, reponse, idNotif):
    service = getGoogleCalendarService()
    event = service.events().get(calendarId='primary', eventId=id).execute()
    for attendees in event["attendees"]:
        if attendees["email"] == current_user.email:
            if reponse == "refuser":
                attendees["responseStatus"] = "declined"
            elif reponse == "valider":
                attendees["responseStatus"] = "accepted"

    service.events().update(calendarId='primary', eventId=id, body=event).execute()
    notification.setStatusById(mysql, idNotif, "accepté");
    return redirect(url_for("showNotifs"))

@app.route('/downloadFile', methods=["GET", "POST"])
def downloadFile():
    print("je download")
    return send_from_directory(directory='static', filename='socialPlanning.apk', as_attachment=True)

# --------lancement de l'application--------
os.environ['DEBUG'] = '1'
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'

# if os.getenv("DATABASE_URL") is not None:
#     port = int(os.environ.get("PORT", 5000))
#     app.run(host='0.0.0.0', port=port)
# else:
#     app.run(ssl_context="adhoc")
if __name__ == '__main__':
    app.run()
